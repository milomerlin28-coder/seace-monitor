import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime
EMAIL_ORIGEN = "milo.merlin28@gmail.com"
EMAIL_CLAVE  = "tqtqpdadifxyxmtx"
EMAIL_DESTINO = "milo.merlin28@gmail.com"

URL = "https://tinyurl.com/conosceconvocatorias2025"

REGIONES = ["CUSCO", "PUNO", "APURIMAC", "MADRE DE DIOS", "AREQUIPA", "AYACUCHO", "MOQUEGUA"]

PALABRAS_CLAVE = [
    "MADERA", "TABLERO", "TRIPLAY", "PARQUET", "MACHIMBRE",
    "PROTECCION PERSONAL", "EPP", "CASCO", "GUANTES", "LENTES DE SEGURIDAD",
    "BOTAS DE SEGURIDAD", "CHALECO", "ARNES", "RESPIRADOR", "MASCARILLA",
    "UNIFORME", "INDUMENTARIA", "ROPA DE TRABAJO", "CALZADO", "POLO",
    "BUZO", "OVEROL", "TERNO", "VESTIMENTA",
    "MEDICAMENTO", "FARMACO", "MEDICINA", "ANTIBIOTICO", "ANALGESICO",
    "INSUMO MEDICO", "MATERIAL MEDICO", "REACTIVO", "VACUNA",
    "EQUIPO DE PROTECCION", "SEGURIDAD INDUSTRIAL", "BOTIQUIN"
]

COLUMNAS = [
    "nroconvocatoria", "descripcion_item", "entidad", "objetocontractual",
    "tipoprocesoseleccion", "sistema_contratacion", "montoreferencial",
    "moneda", "departamento_item", "provincia_item", "distrito_item",
    "fecha_convocatoria", "fechapresentacionpropuesta", "estadoitem"
]

NOMBRES = [
    "N° Convocatoria", "Descripción / Objeto", "Entidad", "Tipo objeto",
    "Tipo de proceso", "Sistema contratación", "Monto referencial",
    "Moneda", "Departamento", "Provincia", "Distrito",
    "Fecha convocatoria", "Fecha presentación", "Estado"
]

print("Descargando datos del OSCE...")
print("Espera 1-2 minutos...")

try:
    respuesta = requests.get(URL, timeout=120, allow_redirects=True)
    print(f"Archivo descargado: {len(respuesta.content)/1024/1024:.1f} MB")

    archivo_osce = openpyxl.load_workbook(BytesIO(respuesta.content), read_only=True)
    hoja_osce = archivo_osce.active
    filas = list(hoja_osce.rows)
    encabezados = [str(c.value).strip() if c.value else "" for c in filas[0]]

    print("Filtrando procesos de tus regiones...")
    procesos = []
    for fila in filas[1:]:
        valores = [c.value for c in fila]
        fila_dict = dict(zip(encabezados, valores))

        region = str(fila_dict.get("departamento_item", "")).upper()
        objeto = str(fila_dict.get("objetocontractual", "")).upper()
        descripcion = str(fila_dict.get("descripcion_item", "")).upper()
        texto = objeto + " " + descripcion

        if any(r in region for r in REGIONES):
            if "BIEN" in objeto and any(kw in texto for kw in PALABRAS_CLAVE):
                procesos.append(fila_dict)

    print(f"Procesos encontrados: {len(procesos)}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Procesos SEACE"

    color_header = PatternFill("solid", fgColor="1F4E79")
    fuente_header = Font(color="FFFFFF", bold=True, size=11)

    for col, nombre in enumerate(NOMBRES, start=1):
        celda = ws.cell(row=1, column=col, value=nombre)
        celda.fill = color_header
        celda.font = fuente_header
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 35

    for fila_num, proceso in enumerate(procesos, start=2):
        for col, campo in enumerate(COLUMNAS, start=1):
            ws.cell(row=fila_num, column=col, value=proceso.get(campo, ""))

    anchos = [15, 45, 40, 12, 25, 20, 15, 8, 15, 15, 15, 18, 18, 12]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    fecha_hoy = datetime.now().strftime("%Y-%m-%d")
    nombre_archivo = f"SEACE_Bienes_{fecha_hoy}.xlsx"
    wb.save(nombre_archivo)

    print(f"\n✓ Archivo Excel creado: {nombre_archivo}")
    print(f"✓ Total de procesos guardados: {len(procesos)}")
    print(f"✓ Ubicación: C:\\Users\\pc\\Desktop\\{nombre_archivo}")
    # Enviar email con Excel adjunto
    # Enviar email con Excel adjunto
    import yagmail

    print("\nEnviando email...")
    yag = yagmail.SMTP(EMAIL_ORIGEN, EMAIL_CLAVE)
    yag.send(
        to=EMAIL_DESTINO,
        subject=f"SEACE Monitor — {len(procesos)} procesos de bienes — {fecha_hoy}",
        contents=f"Buenos días,\n\nTu sistema encontró {len(procesos)} procesos de bienes.\n\nRegiones: Cusco, Puno, Apurímac, Madre de Dios, Arequipa, Ayacucho, Moquegua\n\nSe adjunta el reporte Excel.",
        attachments=nombre_archivo
    )
    print(f"✓ Email enviado a: {EMAIL_DESTINO}")

except Exception as e:
    print(f"Error: {e}")